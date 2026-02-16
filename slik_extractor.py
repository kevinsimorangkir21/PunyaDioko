"""
Extract SLIK OJK (Sistem Layanan Informasi Keuangan) - Kredit/Pembiayaan data
from PDF to Excel/CSV.

Usage:
    python extract_slik.py <input.pdf> [output.xlsx]

If output filename not specified, defaults to slik_output.xlsx
"""

import sys
import re
import os
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


# ── terminal progress bar (no external deps) ─────────────────────────────────

BAR_WIDTH = 30
CLR = "\033[K"   # clear to end of line

def _bar(done: int, total: int) -> str:
    pct    = done / total if total else 1
    filled = int(BAR_WIDTH * pct)
    return "█" * filled + "░" * (BAR_WIDTH - filled)

def print_progress(done: int, total: int, label: str = "") -> None:
    suffix = f"  {label}" if label else ""
    pct    = int((done / total * 100) if total else 100)
    print(f"\r  [{_bar(done, total)}] {pct:3d}%  ({done}/{total} hal){suffix}{CLR}",
          end="", flush=True)


# ── helpers ──────────────────────────────────────────────────────────────────

def clean(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def extract_field(text: str, label: str, terminators: list[str] | None = None) -> str:
    """
    Extract the value that follows `label` in `text`.
    Stops at any of `terminators` (if given).
    """
    pattern = re.escape(label) + r"\s*(.*)"
    m = re.search(pattern, text, re.IGNORECASE)
    if not m:
        return ""
    value = m.group(1).strip()
    if terminators:
        for t in terminators:
            idx = value.lower().find(t.lower())
            if idx >= 0:
                value = value[:idx]
    return clean(value)


def extract_rp(text: str, label: str) -> str:
    """Extract 'Rp X,XX' value after label."""
    pattern = re.escape(label) + r"\s*Rp\s*([\d.,]+)"
    m = re.search(pattern, text, re.IGNORECASE)
    if not m:
        return ""
    return "Rp " + m.group(1).strip()


# ── debitur info ──────────────────────────────────────────────────────────────

def extract_debitur_name(full_text: str) -> str:
    """
    Get debtor name from 'Nama Sesuai Identitas' row.
    Falls back to the top-of-report 'Nama' field.
    """
    m = re.search(r"Nama Sesuai Identitas\s+Identitas.*?\n(\S[^\n]+?)\s+NIK", full_text, re.DOTALL)
    if m:
        return clean(m.group(1))
    m = re.search(r"Nama\s+Jenis Kelamin\s+\d+\n(\S[^\n]+)", full_text)
    if m:
        return clean(m.group(1))
    m = re.search(r"Nama\s*\n([A-Z ]+)", full_text)
    if m:
        return clean(m.group(1))
    return ""


def extract_nomor_laporan(full_text: str) -> str:
    # Pattern: "Nomor Laporan\nNama ... 41897/IDEB/0101564/2019"
    m = re.search(r"(\d+/IDEB/[\d/]+)", full_text)
    return m.group(1).strip() if m else ""


# ── kredit block extraction ───────────────────────────────────────────────────

def split_bank_cabang(pelapor_full: str) -> tuple[str, str]:
    """
    The PDF puts Pelapor and Cabang on the same line and the bank name appears twice.
    E.g.: 'BANK MANDIRI BANK MANDIRI KC TJ.PINANG'
    We detect the shortest repeating prefix to isolate bank name from cabang.
    """
    words = pelapor_full.split()
    # Try to find a repeating prefix (must have non-empty branch remainder)
    for i in range(2, len(words) + 1):
        candidate = " ".join(words[:i])
        rest = pelapor_full[len(candidate):].strip()
        if rest.startswith(candidate):
            cabang = rest[len(candidate):].strip()
            # Only accept if there's a real branch suffix, not just an empty string
            if cabang:
                return candidate, cabang
    # Fallback: split on branch office marker (KC, KCP, KCK, CAPEM, etc.)
    # The string may still have the bank name doubled: 'BCA BCA KCP SUDIRMAN'
    # Strip the repeated first word(s) before the KC marker to get clean bank name.
    m = re.search(r"\s+(KC[A-Z]*|CAPEM|KANTOR\s+CABANG)(?:\s|$)", pelapor_full, re.IGNORECASE)
    if m:
        before_kc = pelapor_full[:m.start()].strip()  # e.g. 'BCA BCA' or 'BANK MANDIRI BANK MANDIRI'
        # Halve by finding repeating prefix in the before_kc part
        bwords = before_kc.split()
        half = len(bwords) // 2
        if half > 0:
            first_half = " ".join(bwords[:half])
            second_half = " ".join(bwords[half:])
            if first_half == second_half:
                before_kc = first_half
        return before_kc, pelapor_full[m.start():].strip()
    return pelapor_full, ""


def extract_credit_blocks(full_text: str) -> list[dict]:
    """
    Extract all Kredit/Pembiayaan (and other facility) blocks from the full text.

    Strategy:
      1. Find every bank-block header (NNN - BANKNAME ... Rp X DATE) by position.
      2. Find every facility-section heading (Garansi Yang Diberikan, etc.) by position.
      3. For each bank-block, the facility type = the last section heading that appears
         *before* that block's start position (and after the previous block).
         If no such heading exists, it's a regular Kredit/Pembiayaan.
      4. Extract all field values from the text slice belonging to each block.
    """
    # ── 1. locate all bank-block headers ────────────────────────────────────
    header_re = re.compile(
        r"(\d{3})\s*-\s*(.+?)\s+Rp\s*([\d.,]+)\s+(\d{2}\s+\w+\s+\d{4})",
        re.MULTILINE,
    )
    bank_headers = list(header_re.finditer(full_text))
    if not bank_headers:
        return []

    # ── 2. locate all facility-section headings ──────────────────────────────
    FACILITY_SECTIONS = [
        "Garansi Yang Diberikan",
        "Irrevocable L/C",
        "Surat Berharga",
        "Fasilitas Lain",
    ]
    heading_re = re.compile(
        r"^(" + "|".join(re.escape(s) for s in FACILITY_SECTIONS) + r")\s*$",
        re.IGNORECASE | re.MULTILINE,
    )
    # list of (position, label)
    headings = [(m.start(), m.group(1)) for m in heading_re.finditer(full_text)]

    # ── 3. build text slices and assign facility types ───────────────────────
    records = []
    for i, hm in enumerate(bank_headers):
        block_start = hm.start()
        block_end   = bank_headers[i + 1].start() if i + 1 < len(bank_headers) else len(full_text)
        chunk       = full_text[block_start:block_end]

        # Previous block's end (= current block's start)
        prev_end = bank_headers[i - 1].start() if i > 0 else 0

        # Nearest facility heading that falls between prev_end and block_start
        facility_type = None
        for hpos, hlabel in headings:
            if prev_end <= hpos < block_start:
                facility_type = hlabel   # keep updating → last one wins (closest)

        # ── field helpers ────────────────────────────────────────────────────
        def field(label, terms=None):
            return extract_field(chunk, label, terms)

        # bank name (strip code prefix)
        bank_name, _ = split_bank_cabang(hm.group(2).strip())
        baki_debet   = "Rp " + hm.group(3)

        # Kualitas
        kualitas_m = re.search(r"No Rekening.*?Kualitas\s+(\d+\s*-\s*[^\n]+)", chunk, re.DOTALL)
        if not kualitas_m:
            kualitas_m = re.search(r"Kualitas\s+(\d+\s*-\s*\w[^\n]{0,30})", chunk)
        kualitas = clean(kualitas_m.group(1)) if kualitas_m else ""

        # Jenis Penggunaan
        if facility_type:
            jenis_penggunaan = facility_type
        else:
            jenis_m = re.search(
                r"Jenis Penggunaan\s+([A-Za-z][^\n\d]+?)(?=\s+Frekuensi|\s+Nilai|\s+Suku|\s*\n)",
                chunk,
            )
            jenis_penggunaan = (
                clean(jenis_m.group(1)) if jenis_m
                else field("Jenis Penggunaan", ["Frekuensi", "\n"])
            )

        # Plafon Awal
        plafon_awal = extract_rp(chunk, "Plafon Awal")

        # Suku Bunga (decimal dot → comma)
        suku_m = re.search(r"Suku Bunga/Imbalan\s+([\d.,]+\s*%)", chunk)
        suku_bunga = re.sub(r"(\d)\.(\d)", r"\1,\2", suku_m.group(1).strip()) if suku_m else ""

        # Tanggal Akad Awal
        tgl_akad_m    = re.search(r"Tanggal Akad Awal\s+(\d{2}\s+\w+\s+\d{4})", chunk)
        tgl_akad_awal = tgl_akad_m.group(1).strip() if tgl_akad_m else ""

        # Tanggal Jatuh Tempo
        tgl_jt_m        = re.search(r"Tanggal Jatuh Tempo\s+(\d{2}\s+\w+\s+\d{4})", chunk)
        tgl_jatuh_tempo = tgl_jt_m.group(1).strip() if tgl_jt_m else ""

        # Frekuensi Restrukturisasi: int for kredit, None for facility rows
        if facility_type:
            frekuensi_restr = None
        else:
            frekuensi_m     = re.search(r"Frekuensi Restrukturisasi\s+(\d+)", chunk)
            frekuensi_restr = int(frekuensi_m.group(1)) if frekuensi_m else 0

        records.append({
            "Bank (Pelapor)":            bank_name,
            "Jenis Penggunaan":          jenis_penggunaan,
            "Nomor Laporan":             "",          # filled from global below
            "Plafon Awal":               plafon_awal,
            "Baki Debet":                baki_debet,
            "Suku Bunga/Imbalan":        suku_bunga,
            "Tanggal Akad Awal":         tgl_akad_awal,
            "Tanggal Jatuh Tempo":       tgl_jatuh_tempo,
            "Kualitas":                  kualitas,
            "Frekuensi Restrukturisasi": frekuensi_restr,
        })

    return records


# ── xlsx writer ───────────────────────────────────────────────────────────────

COLUMNS = [
    "Bank (Pelapor)",
    "Jenis Penggunaan",
    "Nomor Laporan",
    "Plafon Awal",
    "Baki Debet",
    "Suku Bunga/Imbalan",
    "Tanggal Akad Awal",
    "Tanggal Jatuh Tempo",
    "Kualitas",
    "Frekuensi Restrukturisasi",
]

HEADER_BG   = "1F3864"   # dark navy
HEADER_FONT = "FFFFFF"   # white
ALT_ROW_BG  = "D9E1F2"   # light blue

thin = Side(style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def write_xlsx(records: list[dict], debitur: str, out_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kredit_Pembiayaan"

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    title_cell = ws.cell(row=1, column=1,
        value=f"SLIK OJK – Kredit/Pembiayaan  |  Debitur: {debitur}  |  Diekstrak: {datetime.now():%d %B %Y}")
    title_cell.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    title_cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Header row
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font      = Font(name="Arial", bold=True, color=HEADER_FONT, size=10)
        cell.fill      = PatternFill("solid", fgColor="2E5090")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[2].height = 30

    # Data rows
    for row_idx, rec in enumerate(records, start=3):
        bg = ALT_ROW_BG if row_idx % 2 == 0 else "FFFFFF"
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            raw_val = rec.get(col_name, "")
            # None → empty cell (used for Frekuensi on non-kredit rows)
            cell_val = "" if raw_val is None else raw_val
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_val)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border    = border
        ws.row_dimensions[row_idx].height = 18

    # Column widths (manual, tuned to field content)
    col_widths = [30, 18, 32, 18, 18, 14, 18, 18, 16, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes below header
    ws.freeze_panes = "A3"

    wb.save(out_path)


# ── main ──────────────────────────────────────────────────────────────────────

SEP = "─" * 54

def auto_output_name(pdf_path: str) -> str:
    """Derive output .xlsx name from the input PDF filename."""
    base = os.path.splitext(os.path.basename(pdf_path))[0]
    # Sanitise: keep letters, digits, hyphens, underscores
    safe = re.sub(r"[^\w\-]", "_", base).strip("_")
    safe = re.sub(r"_+", "_", safe)          # collapse runs
    return f"{safe}_slik.xlsx"


def process(pdf_path: str, out_path: str) -> None:
    t0 = __import__("time").time()

    print(f"\n{SEP}")
    print(f"  SLIK OJK Extractor")
    print(SEP)
    print(f"  Input  : {os.path.basename(pdf_path)}")
    print(f"  Output : {out_path}")
    print(SEP)

    # ── page-by-page reading with progress bar ────────────────────────────
    pages_text: list[str] = []
    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)
        print(f"\n  Membaca {total} halaman PDF…")
        print_progress(0, total)

        for i, page in enumerate(pdf.pages, 1):
            text = page.extract_text(x_tolerance=3, y_tolerance=3) or ""
            pages_text.append(text)

            # Peek: does this page have a credit block?
            has_kredit = bool(re.search(r"\d{3}\s*-\s*[A-Z]", text))
            label = "kredit ditemukan ✓" if has_kredit else ""
            print_progress(i, total, label)

    print()   # newline after bar

    # ── parse ─────────────────────────────────────────────────────────────
    full_text     = "\n".join(pages_text)
    debitur       = extract_debitur_name(full_text)
    nomor_laporan = extract_nomor_laporan(full_text)

    print(f"\n  Debitur       : {debitur or '(tidak ditemukan)'}")
    print(f"  Nomor Laporan : {nomor_laporan or '(tidak ditemukan)'}")

    print(f"\n  Mengekstrak blok Kredit/Pembiayaan…")
    records = extract_credit_blocks(full_text)

    if not records:
        print("\n  ⚠️  Tidak ditemukan blok Kredit/Pembiayaan.")
        return

    for r in records:
        r["Nomor Laporan"] = nomor_laporan

    print(f"  {len(records)} kredit/pembiayaan ditemukan\n")
    print(f"  {'#':<4} {'Bank (Pelapor)':<24} {'Kualitas':<14} {'Plafon Awal':>16}  {'Baki Debet':>16}")
    print(f"  {'─'*4} {'─'*24} {'─'*14} {'─'*16}  {'─'*16}")
    for i, r in enumerate(records, 1):
        bank  = r["Bank (Pelapor)"][:24]
        kual  = r["Kualitas"][:14]
        plafon = r["Plafon Awal"]
        baki   = r["Baki Debet"]
        print(f"  {i:<4} {bank:<24} {kual:<14} {plafon:>16}  {baki:>16}")

    # ── write ─────────────────────────────────────────────────────────────
    print(f"\n  Menyimpan file Excel…")
    write_xlsx(records, debitur, out_path)

    elapsed = __import__("time").time() - t0
    print(f"\n{SEP}")
    print(f"  ✅ Selesai dalam {elapsed:.1f}s  →  {out_path}")
    print(f"{SEP}\n")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_slik.py <input.pdf> [output.xlsx]")
        sys.exit(1)

    pdf_in   = sys.argv[1]
    xlsx_out = sys.argv[2] if len(sys.argv) > 2 else auto_output_name(pdf_in)
    process(pdf_in, xlsx_out)